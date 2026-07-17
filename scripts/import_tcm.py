import os
import sys
import re
import glob
import json
import urllib.request
import openpyxl

# Setup relative paths
script_dir = os.path.dirname(os.path.abspath(__file__))
project_dir = os.path.dirname(script_dir)
parent_dir = os.path.dirname(project_dir)
env_file_path = os.path.join(project_dir, ".env")

def clean_value(val):
    if not val:
        return ""
    val = val.strip()
    # Remove XML tags like <issue id="...">NUH-1199</issue>
    val = re.sub(r'<issue id=".*?">(.*?)</issue>', r'\1', val)
    # Replace * with empty string
    if val == "*":
        return ""
    return val

def load_env(filepath):
    env_vars = {}
    if not os.path.exists(filepath):
        return env_vars
    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            if '=' in line:
                key, val = line.split('=', 1)
                env_vars[key.strip()] = val.strip().strip('"').strip("'")
    return env_vars

def parse_markdown_tables(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Split content by tables
    lines = content.split('\n')
    tables = []
    current_table = []
    
    for line in lines:
        if line.strip().startswith('|'):
            # Skip separator line like |--|--|
            if re.match(r'^\|\s*[-:]+\s*\|', line.strip()) or '--' in line:
                continue
            # Parse row
            row = [clean_value(cell) for cell in line.split('|')[1:-1]]
            current_table.append(row)
        else:
            if current_table:
                tables.append(current_table)
                current_table = []
    if current_table:
        tables.append(current_table)
        
    return tables

def find_excel_file(site_name):
    # Search patterns:
    # 1. TCM_HIS_Cortex_*_{SITE}.xlsx in parent directory
    # 2. TCM_HIS_Cortex_*.xlsx in parent directory (Fallback)
    search_dir = parent_dir
    
    if site_name:
        pattern = os.path.join(search_dir, f"TCM_HIS_Cortex_*_{site_name}.xlsx")
        files = glob.glob(pattern)
        if files:
            # Return the latest one if multiple found
            return sorted(files)[-1]
            
    # Fallback to general file
    pattern = os.path.join(search_dir, "TCM_HIS_Cortex_*.xlsx")
    files = glob.glob(pattern)
    # Filter out files with specific site names if we just want the master file
    if files:
        # Prefer file without site suffix if available
        for f in files:
            filename = os.path.basename(f)
            # Check if it doesn't end with _NUH, _TMH, _SBH
            if not any(suffix in filename for suffix in ["_NUH.xlsx", "_TMH.xlsx", "_SBH.xlsx"]):
                return f
        return sorted(files)[-1]
        
    return None

def post_comment_to_linear(api_key, issue_identifier, comment_body):
    url = "https://api.linear.app/graphql"
    headers = {
        "Content-Type": "application/json",
        "Authorization": api_key
    }
    
    # Step 1: Query for the UUID of the issue
    query_id = """
    query GetIssueId($id: String!) {
      issue(id: $id) {
        id
      }
    }
    """
    data_id = {
        "query": query_id,
        "variables": {"id": issue_identifier}
    }
    
    req_id = urllib.request.Request(
        url,
        data=json.dumps(data_id).encode("utf-8"),
        headers=headers,
        method="POST"
    )
    
    try:
        with urllib.request.urlopen(req_id) as response:
            res = json.loads(response.read().decode("utf-8"))
            issue_uuid = res.get("data", {}).get("issue", {}).get("id")
            if not issue_uuid:
                print(f"Warning: Could not resolve Linear issue UUID for {issue_identifier}. Comment not posted.")
                return False
    except Exception as e:
        print(f"Warning: Failed to fetch issue UUID: {str(e)}")
        return False
        
    # Step 2: Post the comment using UUID
    query_comment = """
    mutation CreateComment($issueId: String!, $body: String!) {
      commentCreate(input: { issueId: $issueId, body: $body }) {
        success
        comment {
          id
        }
      }
    }
    """
    data_comment = {
        "query": query_comment,
        "variables": {
            "issueId": issue_uuid,
            "body": comment_body
        }
    }
    
    req_comment = urllib.request.Request(
        url,
        data=json.dumps(data_comment).encode("utf-8"),
        headers=headers,
        method="POST"
    )
    
    try:
        with urllib.request.urlopen(req_comment) as response:
            res_comment = json.loads(response.read().decode("utf-8"))
            success = res_comment.get("data", {}).get("commentCreate", {}).get("success")
            if success:
                print(f"Successfully posted reference comment to Linear issue {issue_identifier}!")
                return True
            else:
                print("Warning: Linear API returned unsuccessful comment creation.")
                print(json.dumps(res_comment, indent=2))
                return False
    except Exception as e:
        print(f"Warning: Failed to post comment: {str(e)}")
        return False

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 import_tcm.py <Markdown-File-Path>")
        print("Example: python3 import_tcm.py ../workbooks/md\\ from\\ linear/NUH-1199-test-cases.md")
        sys.exit(1)
        
    input_md_path = sys.argv[1]
    if not os.path.exists(input_md_path):
        print(f"Error: Input Markdown file not found at: {input_md_path}")
        sys.exit(1)
        
    # Read environment config
    print(f"Reading configuration from {env_file_path}...")
    env = load_env(env_file_path)
    site_name = env.get("CURRENT_SITE", "").upper()
    api_key = env.get("LINEAR_API_TOKEN")
    
    print(f"Current Site Configured: {site_name if site_name else 'None (Using Fallback)'}")
    
    # Auto-discover target Excel file
    excel_file_path = env.get("TARGET_EXCEL")
    if excel_file_path:
        if not os.path.isabs(excel_file_path):
            excel_file_path = os.path.join(project_dir, excel_file_path)
        print(f"Using Excel file from config (TARGET_EXCEL): {excel_file_path}")
    else:
        excel_file_path = find_excel_file(site_name)
        
    if not excel_file_path:
        print("Error: Could not discover any TCM_HIS_Cortex_*.xlsx file.")
        sys.exit(1)
        
    excel_filename = os.path.basename(excel_file_path)
    print(f"Target Excel file discovered: {excel_file_path}")
    
    # Parse Markdown Tables
    print(f"Parsing Markdown file {input_md_path}...")
    tables = parse_markdown_tables(input_md_path)
    
    if len(tables) < 2:
        print("Error: Expected at least 2 tables (Scenarios and Test Cases) in the Markdown file.")
        sys.exit(1)
        
    scenarios_data = tables[0]
    testcases_data = tables[1]
    
    # Remove header rows
    scenarios_headers = scenarios_data.pop(0)
    testcases_headers = testcases_data.pop(0)
    
    print(f"Found {len(scenarios_data)} scenarios and {len(testcases_data)} test cases to import.")
    
    # Deduce Card ID from Scenario ID or filename
    # Example: TS-NUH-1199-01 -> Card ID: NUH-1199
    card_id = None
    if scenarios_data:
        first_scen_id = scenarios_data[0][0]
        # Match pattern like TS-NUH-1199 or TS-NUH-1199-01
        match = re.search(r'TS-([A-Za-z0-9]+-\d+)', first_scen_id)
        if match:
            card_id = match.group(1)
            
    if not card_id:
        # Fallback to parse from filename
        basename = os.path.basename(input_md_path)
        match = re.search(r'([A-Za-z0-9]+-\d+)', basename)
        if match:
            card_id = match.group(1)
            
    if not card_id:
        print("Warning: Could not deduce Card ID (e.g. NUH-1199) from Markdown content or filename.")
        
    print(f"Deduced Card ID: {card_id}")
    
    # Load Excel Workbook (preserving formulas)
    print("Loading Excel file...")
    wb = openpyxl.load_workbook(excel_file_path, data_only=False)
    
    # ---------------------------------------------
    # 1. Clean and Update Scenarios Sheet
    # ---------------------------------------------
    ws_scenarios = wb["2. Scenarios"]
    if card_id:
        print(f"Cleaning existing scenarios for Card ID {card_id}...")
        i = ws_scenarios.max_row
        deleted_scen_count = 0
        while i > 1:
            scen_id = ws_scenarios.cell(row=i, column=1).value
            c_id = ws_scenarios.cell(row=i, column=8).value
            if (scen_id and card_id in str(scen_id)) or (c_id and str(c_id) == card_id):
                print(f"  Deleting Scenario at row {i}: {scen_id}")
                ws_scenarios.delete_rows(i)
                deleted_scen_count += 1
            i -= 1
        print(f"Deleted {deleted_scen_count} old scenarios.")
        
    # Re-calculate max row in Scenarios
    last_row_scen = ws_scenarios.max_row
    while last_row_scen > 1 and not ws_scenarios.cell(row=last_row_scen, column=1).value:
        last_row_scen -= 1
        
    # Append scenarios
    new_scen_count = 0
    for scen in scenarios_data:
        new_row = last_row_scen + 1
        print(f"Adding new scenario '{scen[0]}' at row {new_row}")
        for col_idx, val in enumerate(scen, start=1):
            ws_scenarios.cell(row=new_row, column=col_idx, value=val)
        last_row_scen = new_row
        new_scen_count += 1
        
    # ---------------------------------------------
    # 2. Clean and Update Master_TCM Sheet
    # ---------------------------------------------
    ws_master = wb["1. Master_TCM"]
    if card_id:
        print(f"Cleaning existing test cases for Card ID {card_id}...")
        j = ws_master.max_row
        deleted_tc_count = 0
        while j > 1:
            tc_id = ws_master.cell(row=j, column=1).value
            card_ref = ws_master.cell(row=j, column=22).value  # Column V (Cards Requirment)
            if (tc_id and card_id in str(tc_id)) or (card_ref and card_id in str(card_ref)):
                print(f"  Deleting Test Case at row {j}: {tc_id}")
                ws_master.delete_rows(j)
                deleted_tc_count += 1
            j -= 1
        print(f"Deleted {deleted_tc_count} old test cases.")
        
    # Re-calculate max row in Master_TCM
    last_row_tcm = ws_master.max_row
    while last_row_tcm > 1 and not ws_master.cell(row=last_row_tcm, column=1).value:
        last_row_tcm -= 1
        
    target_row = last_row_tcm + 1
    start_written_row = target_row
    
    # Precise Markdown to Excel column mapping
    md_excel_map = [
        (0, 1),   # TC ID -> A
        (1, 2),   # Scenario ID -> B
        (2, 3),   # Module -> C
        (3, 5),   # Priority -> E    (D is Feature formula)
        (4, 6),   # Test Case Name -> F
        (5, 7),   # Pre-conditions -> G
        (6, 8),   # Test Steps -> H
        (7, 9),   # Test Data -> I
        (8, 10),  # BDD -> J
        (9, 11),  # Expected Result -> K
        (10, 12), # Test Level -> L
        (11, 13), # Test Type -> M
        (12, 14), # Case Type -> N
        (13, 15), # Create By -> O   (P is Create Date formula)
        (14, 21), # Execution Status -> U
        (15, 22), # Cards Requirment -> V
        (16, 23), # Requirment Detail -> W
        (17, 24), # Run in This Cycle? -> X
        (18, 25)  # Remark -> Y
    ]
    
    new_tc_count = 0
    for tc in testcases_data:
        tc_id = tc[0]
        print(f"Writing Test Case '{tc_id}' to row {target_row}")
        
        # Write values
        for tc_idx, col_idx in md_excel_map:
            if tc_idx < len(tc):
                val = tc[tc_idx]
                ws_master.cell(row=target_row, column=col_idx, value=val)
                
        # Write formulas for D, P, R, T
        ws_master.cell(row=target_row, column=4, value=f'="IPD - "&C{target_row}')
        ws_master.cell(row=target_row, column=16, value=f'=IF(O{target_row}<>"",IF(OR(P{target_row}="",P{target_row}=0),NOW(),P{target_row}),"")')
        ws_master.cell(row=target_row, column=18, value=f'=IF(Q{target_row}<>"",IF(OR(R{target_row}="",R{target_row}=0),NOW(),R{target_row}),"")')
        ws_master.cell(row=target_row, column=20, value=f'=IF(S{target_row}<>"",IF(OR(T{target_row}="",T{target_row}=0),NOW(),T{target_row}),"")')
        
        # Unhide newly written row
        ws_master.row_dimensions[target_row].hidden = False
        
        last_row_tcm = max(last_row_tcm, target_row)
        target_row += 1
        new_tc_count += 1
        
    end_written_row = target_row - 1
    
    # Update table references
    try:
        if "Master_TCM" in ws_master.tables:
            table = ws_master.tables["Master_TCM"]
            table.ref = f"A1:Y{last_row_tcm}"
            print(f"Updated Table 'Master_TCM' range to {table.ref}")
    except Exception as e:
        print("Warning: Could not update Master_TCM table reference:", str(e))
        
    try:
        if "Scenarios" in ws_scenarios.tables:
            table = ws_scenarios.tables["Scenarios"]
            table.ref = f"A1:H{last_row_scen}"
            print(f"Updated Table 'Scenarios' range to {table.ref}")
    except Exception as e:
        print("Warning: Could not update Scenarios table reference:", str(e))
        
    # Save modified Workbook
    print("Saving modified workbook...")
    wb.save(excel_file_path)
    print(f"Success! Imported scenarios and test cases into {excel_file_path}")
    
    # ---------------------------------------------
    # 3. Post reference comment back to Linear Card
    # ---------------------------------------------
    if api_key and card_id and new_tc_count > 0:
        print(f"Posting documentation reference comment back to Linear card {card_id}...")
        
        # Build comment message in English/Thai as standard
        comment_body = f"""📝 **Test Cases Linked Successfully**
* **Excel Workbook:** `{excel_filename}` (Sheet: `1. Master_TCM`, Rows: **{start_written_row}-{end_written_row}**)
* **Markdown Draft Source:** `{os.path.basename(input_md_path)}`
* **Total Test Cases:** {new_tc_count} cases
* **Execution Status:** Draft

*This is an automated reference post by RTM Bridge Checker.*"""
        
        post_comment_to_linear(api_key, card_id, comment_body)
    else:
        print("Skipped comment posting (missing API token, Card ID, or written cases).")

if __name__ == "__main__":
    main()
